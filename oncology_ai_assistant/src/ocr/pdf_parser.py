"""
=============================================================================
PDF_PARSER.PY - Парсер PDF документов и извлечение текста
=============================================================================
Модуль для работы с PDF документами:
- Извлечение текста из PDF (включая сканы)
- Конвертация страниц PDF в изображения
- Обработка таблиц и структурированных данных
- Поддержка медицинских форматов (выписки, заключения)
=============================================================================
"""

import logging
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import tempfile
import os

from PIL import Image

# pdfplumber для текстовых PDF
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

# PyMuPDF (fitz) для конвертации в изображения
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False


logger = logging.getLogger(__name__)


@dataclass
class PDFPage:
    """Страница PDF с содержимым."""
    page_number: int
    text: str
    tables: List[List[List[str]]] = field(default_factory=list)
    images_count: int = 0
    width: float = 0.0
    height: float = 0.0
    
    def to_dict(self) -> Dict[str, Any]:
        """Сериализация в словарь."""
        return {
            'page_number': self.page_number,
            'text': self.text,
            'tables_count': len(self.tables),
            'tables': self.tables,
            'images_count': self.images_count,
            'dimensions': (self.width, self.height)
        }


@dataclass
class PDFDocument:
    """Представление PDF документа."""
    path: str
    total_pages: int
    pages: List[PDFPage] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    is_scanned: bool = False
    processed_at: datetime = field(default_factory=datetime.now)
    
    @property
    def full_text(self) -> str:
        """Полный текст документа."""
        return '\n'.join(page.text for page in self.pages)
    
    @property
    def has_text(self) -> bool:
        """Есть ли извлекаемый текст."""
        return any(page.text.strip() for page in self.pages)
    
    def to_dict(self) -> Dict[str, Any]:
        """Сериализация в словарь."""
        return {
            'path': self.path,
            'total_pages': self.total_pages,
            'pages': [p.to_dict() for p in self.pages],
            'metadata': self.metadata,
            'is_scanned': self.is_scanned,
            'full_text': self.full_text,
            'processed_at': self.processed_at.isoformat()
        }


class PDFParser:
    """
    Парсер PDF документов с извлечением текста и таблиц.
    
    Использование:
        parser = PDFParser()
        doc = parser.parse('document.pdf')
        print(doc.full_text)
    """
    
    def __init__(
        self,
        extract_tables: bool = True,
        text_tolerance: float = 3.0
    ):
        """
        Инициализация парсера.
        
        Args:
            extract_tables: Извлекать таблицы.
            text_tolerance: Точность позиционирования текста.
        """
        self.extract_tables = extract_tables
        self.text_tolerance = text_tolerance
        
        if not PDFPLUMBER_AVAILABLE:
            logger.warning("pdfplumber не установлен. Текст из PDF не будет извлечён.")
        if not PYMUPDF_AVAILABLE:
            logger.warning("PyMuPDF не установлен. Конвертация в изображения не будет работать.")
    
    def parse(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None
    ) -> PDFDocument:
        """
        Распарсить PDF документ.
        
        Args:
            pdf_path: Путь к PDF файлу.
            pages: Список страниц для обработки (None = все).
            
        Returns:
            PDFDocument с извлечёнными данными.
        """
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF файл не найден: {pdf_path}")
        
        logger.info(f"Парсинг PDF: {pdf_path}")
        
        doc = PDFDocument(
            path=str(pdf_path),
            total_pages=0
        )
        
        # Извлекаем метаданные
        doc.metadata = self._extract_metadata(pdf_path)
        
        # Парсим страницы через pdfplumber
        if PDFPLUMBER_AVAILABLE:
            self._parse_with_pdfplumber(pdf_path, doc, pages)
        
        # Определяем является ли документ сканом
        doc.is_scanned = not doc.has_text
        
        logger.info(
            f"PDF распарсен: {doc.total_pages} страниц, "
            f"скан={doc.is_scanned}, текст={doc.has_text}"
        )
        
        return doc
    
    def _extract_metadata(self, pdf_path: Path) -> Dict[str, Any]:
        """Извлечь метаданные PDF."""
        metadata = {}
        
        try:
            if PYMUPDF_AVAILABLE:
                with fitz.open(pdf_path) as pdf:
                    meta = pdf.metadata
                    metadata = {
                        'title': meta.get('title', ''),
                        'author': meta.get('author', ''),
                        'subject': meta.get('subject', ''),
                        'creator': meta.get('creator', ''),
                        'producer': meta.get('producer', ''),
                        'creation_date': meta.get('creationDate', ''),
                        'modification_date': meta.get('modDate', '')
                    }
        except Exception as e:
            logger.warning(f"Не удалось извлечь метаданные: {e}")
        
        return metadata
    
    def _parse_with_pdfplumber(
        self,
        pdf_path: Path,
        doc: PDFDocument,
        pages: Optional[List[int]]
    ) -> None:
        """Парсинг с помощью pdfplumber."""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                doc.total_pages = len(pdf.pages)
                
                # Определяем какие страницы обрабатывать
                page_indices = pages if pages else range(len(pdf.pages))
                page_indices = [i for i in page_indices if 0 <= i < len(pdf.pages)]
                
                for page_num in page_indices:
                    page = pdf.pages[page_num]
                    pdf_page = self._parse_page(page, page_num + 1)
                    doc.pages.append(pdf_page)
                    
        except Exception as e:
            logger.error(f"Ошибка при парсинге PDF: {e}")
            raise
    
    def _parse_page(self, page, page_number: int) -> PDFPage:
        """
        Распарсить одну страницу.
        
        Args:
            page: Страница pdfplumber.
            page_number: Номер страницы (1-based).
            
        Returns:
            PDFPage с данными.
        """
        # Извлекаем текст
        text = page.extract_text(x_tolerance=self.text_tolerance) or ""
        
        # Извлекаем таблицы
        tables = []
        if self.extract_tables:
            pdf_tables = page.extract_tables()
            for table in pdf_tables:
                if table:
                    # Очищаем таблицу от None
                    cleaned_table = [
                        [cell if cell else "" for cell in row]
                        for row in table
                    ]
                    tables.append(cleaned_table)
        
        # Считаем изображения
        images_count = len(page.images) if hasattr(page, 'images') else 0
        
        # Размеры страницы
        width = page.width if hasattr(page, 'width') else 0
        height = page.height if hasattr(page, 'height') else 0
        
        return PDFPage(
            page_number=page_number,
            text=text.strip(),
            tables=tables,
            images_count=images_count,
            width=width,
            height=height
        )
    
    def pdf_to_images(
        self,
        pdf_path: Union[str, Path],
        dpi: int = 300,
        pages: Optional[List[int]] = None
    ) -> List[Image.Image]:
        """
        Конвертировать PDF страницы в изображения.
        
        Args:
            pdf_path: Путь к PDF.
            dpi: Разрешение изображений.
            pages: Список страниц (None = все).
            
        Returns:
            Список PIL Image.
        """
        if not PYMUPDF_AVAILABLE:
            raise ImportError(
                "PyMuPDF не установлен. Установите: pip install pymupdf"
            )
        
        pdf_path = Path(pdf_path)
        images = []
        
        with fitz.open(pdf_path) as pdf:
            # Определяем страницы
            page_indices = pages if pages else range(len(pdf))
            page_indices = [i for i in page_indices if 0 <= i < len(pdf)]
            
            for page_num in page_indices:
                page = pdf[page_num]
                
                # Матрица масштабирования
                zoom = dpi / 72.0
                matrix = fitz.Matrix(zoom, zoom)
                
                # Рендерим страницу
                pix = page.get_pixmap(matrix=matrix)
                
                # Конвертируем в PIL Image
                img_data = pix.tobytes("png")
                image = Image.open(BytesIO(img_data))
                images.append(image)
        
        logger.info(f"Конвертировано {len(images)} страниц в изображения")
        return images
    
    def extract_tables_from_pdf(
        self,
        pdf_path: Union[str, Path],
        page_numbers: Optional[List[int]] = None
    ) -> Dict[int, List[List[List[str]]]]:
        """
        Извлечь таблицы из PDF.
        
        Args:
            pdf_path: Путь к PDF.
            page_numbers: Номера страниц.
            
        Returns:
            Словарь {номер_страницы: список_таблиц}.
        """
        doc = self.parse(pdf_path, pages=page_numbers)
        
        tables_by_page = {}
        for page in doc.pages:
            if page.tables:
                tables_by_page[page.page_number] = page.tables
        
        return tables_by_page


@dataclass
class ExcelSheet:
    """Лист Excel с данными."""
    name: str
    index: int
    rows: List[List[Any]] = field(default_factory=list)
    headers: Optional[List[str]] = None
    
    @property
    def as_dicts(self) -> List[Dict[str, Any]]:
        """Преобразовать в список словарей (если есть заголовки)."""
        if not self.headers:
            return []
        
        result = []
        for row in self.rows[1:]:  # Пропускаем заголовок
            row_dict = {}
            for i, header in enumerate(self.headers):
                if i < len(row):
                    row_dict[header] = row[i]
                else:
                    row_dict[header] = None
            result.append(row_dict)
        return result


class ExcelParser:
    """Парсер Excel файлов."""
    
    def __init__(self):
        """Инициализация парсера Excel."""
        try:
            import pandas as pd
            self.pd = pd
        except ImportError:
            raise ImportError("pandas не установлен. Установите: pip install pandas")
        
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")
    
    def parse(
        self,
        excel_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None
    ) -> List[ExcelSheet]:
        """
        Распарсить Excel файл.
        
        Args:
            excel_path: Путь к файлу.
            sheet_names: Имена листов для обработки.
            
        Returns:
            Список ExcelSheet.
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel файл не найден: {excel_path}")
        
        sheets = []
        
        # Открываем workbook для получения имён листовов
        wb = self.openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = sheet_names or wb.sheetnames
        
        for idx, sheet_name in enumerate(sheet_names):
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден")
                continue
            
            # Читаем лист через pandas
            df = self.pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            
            # Преобразуем в список списков
            rows = df.values.tolist()
            
            # Пытаемся определить заголовки
            headers = None
            if rows and any(cell is not None for cell in rows[0]):
                headers = [str(cell) if cell is not None else f"Column_{i}" 
                          for i, cell in enumerate(rows[0])]
            
            sheet = ExcelSheet(
                name=sheet_name,
                index=idx,
                rows=rows,
                headers=headers
            )
            sheets.append(sheet)
        
        wb.close()
        logger.info(f"Распарсен Excel: {len(sheets)} листов")
        return sheets
    
    def to_text(
        self,
        excel_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None
    ) -> str:
        """
        Преобразовать Excel в текст.
        
        Args:
            excel_path: Путь к файлу.
            sheet_names: Листы для обработки.
            
        Returns:
            Текстовое представление.
        """
        sheets = self.parse(excel_path, sheet_names)
        
        texts = []
        for sheet in sheets:
            sheet_text = [f"=== ЛИСТ: {sheet.name} ==="]
            
            for row in sheet.rows:
                row_text = [str(cell) if cell is not None else "" for cell in row]
                sheet_text.append(" | ".join(row_text))
            
            texts.append('\n'.join(sheet_text))
        
        return '\n\n'.join(texts)


# -----------------------------------------------------------------------------
# Утилитные функции
# -----------------------------------------------------------------------------

def extract_text_from_pdf(pdf_path: Union[str, Path]) -> str:
    """
    Извлечь текст из PDF.
    
    Args:
        pdf_path: Путь к файлу.
        
    Returns:
        Текст документа.
    """
    parser = PDFParser()
    doc = parser.parse(pdf_path)
    return doc.full_text


def pdf_to_images(
    pdf_path: Union[str, Path],
    dpi: int = 300
) -> List[Image.Image]:
    """
    Конвертировать PDF в изображения.
    
    Args:
        pdf_path: Путь к PDF.
        dpi: Разрешение.
        
    Returns:
        Список изображений.
    """
    parser = PDFParser()
    return parser.pdf_to_images(pdf_path, dpi=dpi)


def extract_text_from_excel(excel_path: Union[str, Path]) -> str:
    """
    Извлечь текст из Excel.
    
    Args:
        excel_path: Путь к файлу.
        
    Returns:
        Текст документа.
    """
    parser = ExcelParser()
    return parser.to_text(excel_path)


# Импорт для BytesIO
from io import BytesIO
